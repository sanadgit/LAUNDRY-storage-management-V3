from pathlib import Path
from datetime import date

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "output" / "spreadsheet"
OUTPUT_FILE = OUTPUT_DIR / "laundry_multi_branch_financial_forecast.xlsx"

NAVY = "17324D"
TEAL = "167D8D"
LIGHT_TEAL = "DCEFF2"
BLUE = "1F4E78"
INPUT_BLUE = "DDEBF7"
GREEN = "E2F0D9"
GRAY = "E7E6E6"
LIGHT_GRAY = "F3F5F7"
ORANGE = "FCE4D6"
YELLOW = "FFF2CC"
RED = "F4CCCC"
PURPLE = "E4DFEC"
WHITE = "FFFFFF"
BLACK = "000000"
DARK_RED = "9C0006"

CURRENCY_FORMAT = '"AED" #,##0.00;[Red]("AED" #,##0.00);-'
NUMBER_FORMAT = '#,##0.00;[Red](#,##0.00);-'
DATE_FORMAT = "dd-mmm-yy"
PERCENT_FORMAT = "0.0%"
THIN_GRAY = Side(style="thin", color="D9E1E8")
MEDIUM_NAVY = Side(style="medium", color=NAVY)


def title(ws, text, subtitle, end_column):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = ws.cell(1, 1, text)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(color=WHITE, bold=True, size=18)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
    cell = ws.cell(2, 1, subtitle)
    cell.fill = PatternFill("solid", fgColor=LIGHT_TEAL)
    cell.font = Font(color=NAVY, italic=True, size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 24


def header_row(ws, row, headers):
    for col, value in enumerate(headers, 1):
        cell = ws.cell(row, col, value)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=MEDIUM_NAVY)
    ws.row_dimensions[row].height = 32


def add_table(ws, name, ref):
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def set_widths(ws, widths):
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width


def style_input_area(ws, min_row, max_row, columns, imported_columns=()):
    imported = set(imported_columns)
    for row in range(min_row, max_row + 1):
        for col in columns:
            cell = ws.cell(row, col)
            cell.fill = PatternFill("solid", fgColor=GREEN if col in imported else INPUT_BLUE)
            cell.font = Font(color="008000" if col in imported else BLUE)
            cell.border = Border(bottom=THIN_GRAY)
            cell.alignment = Alignment(vertical="center", wrap_text=False)


def style_formula_columns(ws, min_row, max_row, columns):
    for row in range(min_row, max_row + 1):
        for col in columns:
            cell = ws.cell(row, col)
            cell.fill = PatternFill("solid", fgColor=WHITE)
            cell.font = Font(color=BLACK)
            cell.border = Border(bottom=THIN_GRAY)


def add_list_validation(ws, cell_range, source_range):
    validation = DataValidation(type="list", formula1=source_range, allow_blank=True)
    validation.error = "Select a value from the list."
    validation.errorTitle = "Invalid value"
    validation.prompt = "Choose from the approved setup list."
    validation.promptTitle = "Controlled input"
    ws.add_data_validation(validation)
    validation.add(cell_range)


def kpi(ws, label_cell, value_cell, label, formula, fill=TEAL, number_format=CURRENCY_FORMAT):
    ws[label_cell] = label
    ws[label_cell].fill = PatternFill("solid", fgColor=NAVY)
    ws[label_cell].font = Font(color=WHITE, bold=True, size=10)
    ws[label_cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[value_cell] = formula
    ws[value_cell].fill = PatternFill("solid", fgColor=fill)
    ws[value_cell].font = Font(color=NAVY, bold=True, size=16)
    ws[value_cell].alignment = Alignment(horizontal="center", vertical="center")
    ws[value_cell].number_format = number_format


def build_setup(wb):
    ws = wb.create_sheet("Setup & Lists")
    title(ws, "Setup & Lists", "قوائم التحكم المركزية — عدّل القيم هنا لتظهر في جميع الصفحات", 12)
    lists = {
        "A": ("Branches", ["MBZ", "Mussafah", "Al Falah"]),
        "B": ("Categories", ["Rent", "Salaries", "Laundry Supplies", "Utilities", "Vehicle Expenses", "Maintenance", "Government Fees", "Marketing", "Other"]),
        "C": ("Payment Methods", ["Cash", "Credit Card", "Bank Transfer", "Cheque", "Credit"]),
        "D": ("Commitment Status", ["Draft", "Pending", "Partially Paid", "Paid", "On Hold", "Cancelled"]),
        "E": ("Cheque Status", ["Pending", "Cleared", "Returned", "Cancelled"]),
        "F": ("Priority", ["Critical", "High", "Normal", "Low"]),
        "G": ("Yes No", ["Yes", "No"]),
        "H": ("Source Types", ["AP", "Cheque", "Fixed", "Manual", "POS", "Payroll"]),
        "I": ("Account Types", ["Bank", "Cash", "Card Clearing"]),
        "J": ("AP Status", ["Open", "Partially Paid", "Paid", "Disputed", "Cancelled"]),
    }
    for col, (heading, values) in lists.items():
        ws[f"{col}4"] = heading
        ws[f"{col}4"].fill = PatternFill("solid", fgColor=TEAL)
        ws[f"{col}4"].font = Font(color=WHITE, bold=True)
        for index, value in enumerate(values, 5):
            ws[f"{col}{index}"] = value
            ws[f"{col}{index}"].fill = PatternFill("solid", fgColor=GRAY)
            ws[f"{col}{index}"].font = Font(color="666666")
    ws["L4"] = "Model Controls"
    ws["L4"].fill = PatternFill("solid", fgColor=PURPLE)
    ws["L4"].font = Font(color=NAVY, bold=True)
    controls = [
        ("L5", "Forecast Months", 12),
        ("L6", "Minimum Liquidity Buffer", 10000),
        ("L7", "Alert Days", 7),
        ("L8", "Currency", "AED"),
        ("L9", "Last POS Sync", "Manual / n8n"),
    ]
    for address, label, value in controls:
        row = ws[address].row
        ws[address] = label
        ws.cell(row, 13, value)
        ws[address].fill = PatternFill("solid", fgColor=GRAY)
        ws.cell(row, 13).fill = PatternFill("solid", fgColor=YELLOW)
        ws.cell(row, 13).font = Font(color=BLUE)
    ws["M6"].number_format = CURRENCY_FORMAT
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    set_widths(ws, {"A": 20, "B": 24, "C": 20, "D": 22, "E": 18, "F": 16, "G": 12, "H": 18, "I": 18, "J": 20, "L": 26, "M": 22})
    return ws


def build_commitments(wb):
    ws = wb.create_sheet("Commitments")
    headers = ["Commitment ID", "Source Type", "Source ID", "Branch", "Due Date", "Category", "Description", "Supplier / Payee", "Amount (AED)", "Payment Method", "Paid Amount (AED)", "Balance (AED)", "Status", "Priority", "Recurring", "Payment Date", "Bank Account", "Approved By", "Notes", "Days to Due", "Alert"]
    title(ws, "Future Commitments — Central Register", "السجل المركزي الوحيد للالتزامات؛ استخدم Source Type + Source ID لمنع العدّ المزدوج", len(headers))
    header_row(ws, 4, headers)
    last_row = 504
    style_input_area(ws, 5, last_row, [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 13, 14, 15, 16, 17, 18, 19], imported_columns=[2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 13])
    style_formula_columns(ws, 5, last_row, [12, 20, 21])
    for row in range(5, last_row + 1):
        ws.cell(row, 12, f'=IF(I{row}="","",MAX(0,I{row}-K{row}))')
        ws.cell(row, 20, f'=IF(E{row}="","",E{row}-TODAY())')
        ws.cell(row, 21, f'=IF(OR(M{row}="Paid",M{row}="Cancelled",E{row}=""),"",IF(E{row}<TODAY(),"OVERDUE",IF(E{row}<=TODAY()+\'Setup & Lists\'!$M$7,"DUE SOON","")))')
        for col in [9, 11, 12]:
            ws.cell(row, col).number_format = CURRENCY_FORMAT
        for col in [5, 16]:
            ws.cell(row, col).number_format = DATE_FORMAT
    add_table(ws, "tblCommitments", f"A4:U{last_row}")
    add_list_validation(ws, f"B5:B{last_row}", "='Setup & Lists'!$H$5:$H$10")
    add_list_validation(ws, f"D5:D{last_row}", "='Setup & Lists'!$A$5:$A$20")
    add_list_validation(ws, f"F5:F{last_row}", "='Setup & Lists'!$B$5:$B$20")
    add_list_validation(ws, f"J5:J{last_row}", "='Setup & Lists'!$C$5:$C$15")
    add_list_validation(ws, f"M5:M{last_row}", "='Setup & Lists'!$D$5:$D$15")
    add_list_validation(ws, f"N5:N{last_row}", "='Setup & Lists'!$F$5:$F$10")
    add_list_validation(ws, f"O5:O{last_row}", "='Setup & Lists'!$G$5:$G$6")
    ws.conditional_formatting.add(f"U5:U{last_row}", FormulaRule(formula=['$U5="OVERDUE"'], fill=PatternFill("solid", fgColor=RED), font=Font(color=DARK_RED, bold=True)))
    ws.conditional_formatting.add(f"U5:U{last_row}", FormulaRule(formula=['$U5="DUE SOON"'], fill=PatternFill("solid", fgColor=ORANGE), font=Font(color="9C5700", bold=True)))
    ws.conditional_formatting.add(f"B5:C{last_row}", FormulaRule(formula=[f'AND($C5<>"",COUNTIFS($B$5:$B${last_row},$B5,$C$5:$C${last_row},$C5)>1)'], fill=PatternFill("solid", fgColor=RED), font=Font(color=DARK_RED)))
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:U{last_row}"
    ws.sheet_view.showGridLines = False
    set_widths(ws, {"A": 18, "B": 14, "C": 18, "D": 14, "E": 13, "F": 20, "G": 30, "H": 24, "I": 16, "J": 18, "K": 18, "L": 17, "M": 18, "N": 12, "O": 12, "P": 14, "Q": 22, "R": 18, "S": 30, "T": 13, "U": 14})
    return ws


def build_cheques(wb):
    ws = wb.create_sheet("Cheque Register")
    headers = ["Cheque No", "Commitment ID", "Branch", "Bank Account", "Beneficiary", "Issue Date", "Cheque Date", "Amount (AED)", "Status", "Cleared Date", "Replacement Cheque No", "Return Reason", "Days to Due", "Alert", "Notes"]
    title(ws, "Cheque Register", "يرتبط كل شيك بالالتزام الأصلي؛ قيمة الشيك لا تُضاف مرة ثانية إلى الالتزامات", len(headers))
    header_row(ws, 4, headers)
    last_row = 304
    style_input_area(ws, 5, last_row, [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 15])
    style_formula_columns(ws, 5, last_row, [13, 14])
    for row in range(5, last_row + 1):
        ws.cell(row, 13, f'=IFERROR(IF(G{row}="","",G{row}-TODAY()),"")')
        ws.cell(row, 14, f'=IFERROR(IF(OR(I{row}="Cleared",I{row}="Cancelled",G{row}=""),"",IF(G{row}<TODAY(),"OVERDUE",IF(G{row}<=TODAY()+\'Setup & Lists\'!$M$7,"DUE SOON",""))),"CHECK DATE")')
        ws.cell(row, 8).number_format = CURRENCY_FORMAT
        for col in [6, 7, 10]:
            ws.cell(row, col).number_format = DATE_FORMAT
    add_table(ws, "tblCheques", f"A4:O{last_row}")
    add_list_validation(ws, f"C5:C{last_row}", "='Setup & Lists'!$A$5:$A$20")
    add_list_validation(ws, f"I5:I{last_row}", "='Setup & Lists'!$E$5:$E$10")
    ws.conditional_formatting.add(f"N5:N{last_row}", FormulaRule(formula=['$N5="OVERDUE"'], fill=PatternFill("solid", fgColor=RED)))
    ws.conditional_formatting.add(f"N5:N{last_row}", FormulaRule(formula=['$N5="DUE SOON"'], fill=PatternFill("solid", fgColor=ORANGE)))
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    set_widths(ws, {"A": 16, "B": 18, "C": 14, "D": 22, "E": 24, "F": 13, "G": 13, "H": 16, "I": 14, "J": 13, "K": 21, "L": 28, "M": 13, "N": 14, "O": 28})
    return ws


def build_ap(wb):
    ws = wb.create_sheet("Accounts Payable")
    headers = ["AP ID", "Commitment ID", "Supplier", "Invoice No", "Invoice Date", "Due Date", "Branch", "Category", "Amount (AED)", "Paid (AED)", "Balance (AED)", "Status", "Payment Method", "Purchase ID", "Duplicate Key", "Overdue Days", "Alert", "Notes"]
    title(ws, "Accounts Payable", "ذمم الموردين؛ كل فاتورة مفتوحة يجب أن ترتبط بـ Commitment ID واحد", len(headers))
    header_row(ws, 4, headers)
    last_row = 504
    style_input_area(ws, 5, last_row, [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 12, 13, 14, 18], imported_columns=[1, 3, 4, 5, 6, 7, 8, 9, 10, 12, 14])
    style_formula_columns(ws, 5, last_row, [11, 15, 16, 17])
    for row in range(5, last_row + 1):
        ws.cell(row, 11, f'=IF(I{row}="","",MAX(0,I{row}-J{row}))')
        ws.cell(row, 15, f'=IF(OR(C{row}="",D{row}=""),"",C{row}&"|"&D{row}&"|"&TEXT(I{row},"0.00"))')
        ws.cell(row, 16, f'=IF(OR(F{row}="",K{row}=0),0,MAX(0,TODAY()-F{row}))')
        ws.cell(row, 17, f'=IF(OR(L{row}="Paid",L{row}="Cancelled",F{row}=""),"",IF(F{row}<TODAY(),"OVERDUE",IF(F{row}<=TODAY()+\'Setup & Lists\'!$M$7,"DUE SOON","")))')
        for col in [9, 10, 11]:
            ws.cell(row, col).number_format = CURRENCY_FORMAT
        for col in [5, 6]:
            ws.cell(row, col).number_format = DATE_FORMAT
    add_table(ws, "tblAP", f"A4:R{last_row}")
    add_list_validation(ws, f"G5:G{last_row}", "='Setup & Lists'!$A$5:$A$20")
    add_list_validation(ws, f"H5:H{last_row}", "='Setup & Lists'!$B$5:$B$20")
    add_list_validation(ws, f"L5:L{last_row}", "='Setup & Lists'!$J$5:$J$12")
    add_list_validation(ws, f"M5:M{last_row}", "='Setup & Lists'!$C$5:$C$15")
    ws.conditional_formatting.add(f"Q5:Q{last_row}", FormulaRule(formula=['$Q5="OVERDUE"'], fill=PatternFill("solid", fgColor=RED)))
    ws.conditional_formatting.add(f"Q5:Q{last_row}", FormulaRule(formula=['$Q5="DUE SOON"'], fill=PatternFill("solid", fgColor=ORANGE)))
    ws.conditional_formatting.add(f"O5:O{last_row}", FormulaRule(formula=[f'AND($O5<>"",COUNTIF($O$5:$O${last_row},$O5)>1)'], fill=PatternFill("solid", fgColor=RED)))
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    set_widths(ws, {"A": 15, "B": 18, "C": 25, "D": 18, "E": 13, "F": 13, "G": 14, "H": 20, "I": 16, "J": 16, "K": 16, "L": 18, "M": 18, "N": 15, "O": 35, "P": 14, "Q": 14, "R": 28})
    return ws


def build_fixed(wb):
    ws = wb.create_sheet("Fixed Expenses")
    headers = ["Fixed ID", "Branch", "Expense Type", "Supplier / Payee", "Monthly Amount (AED)", "Due Day", "Start Date", "End Date", "Payment Method", "Active", "Source Reference", "Last Generated Month", "Next Due Date", "Notes"]
    title(ws, "Fixed Expenses Setup", "قوالب المصروفات الدورية؛ عند توليد الالتزام استخدم Source Type = Fixed وSource ID = Fixed ID", len(headers))
    header_row(ws, 4, headers)
    last_row = 204
    style_input_area(ws, 5, last_row, [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 14])
    style_formula_columns(ws, 5, last_row, [13])
    for row in range(5, last_row + 1):
        ws.cell(row, 13, f'=IF(OR(F{row}="",J{row}<>"Yes"),"",DATE(YEAR(TODAY()),MONTH(TODAY())+(DAY(TODAY())>F{row}),MIN(F{row},DAY(EOMONTH(DATE(YEAR(TODAY()),MONTH(TODAY())+(DAY(TODAY())>F{row}),1),0)))))')
        ws.cell(row, 5).number_format = CURRENCY_FORMAT
        for col in [7, 8, 12, 13]:
            ws.cell(row, col).number_format = DATE_FORMAT
    add_table(ws, "tblFixedExpenses", f"A4:N{last_row}")
    add_list_validation(ws, f"B5:B{last_row}", "='Setup & Lists'!$A$5:$A$20")
    add_list_validation(ws, f"C5:C{last_row}", "='Setup & Lists'!$B$5:$B$20")
    add_list_validation(ws, f"I5:I{last_row}", "='Setup & Lists'!$C$5:$C$15")
    add_list_validation(ws, f"J5:J{last_row}", "='Setup & Lists'!$G$5:$G$6")
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    set_widths(ws, {"A": 15, "B": 14, "C": 24, "D": 24, "E": 20, "F": 12, "G": 13, "H": 13, "I": 18, "J": 11, "K": 20, "L": 20, "M": 15, "N": 30})
    return ws


def build_balances(wb):
    ws = wb.create_sheet("Bank & Cash Balances")
    headers = ["Account ID", "Branch", "Account Name", "Account Type", "As of Date", "Book Balance (AED)", "Pending Cheques (AED)", "Protected Minimum (AED)", "Available Balance (AED)", "Source", "Last Synced At", "Notes"]
    title(ws, "Bank & Cash Balances", "افصل الرصيد الدفتري عن الرصيد المتاح بعد الشيكات والحد الأدنى المحمي", len(headers))
    header_row(ws, 4, headers)
    last_row = 104
    style_input_area(ws, 5, last_row, [1, 2, 3, 4, 5, 6, 8, 10, 11, 12], imported_columns=[1, 2, 3, 4, 5, 6, 10, 11])
    style_formula_columns(ws, 5, last_row, [7, 9])
    starter = [
        ["MBZ-BANK", "MBZ", "Main Bank — MBZ", "Bank"],
        ["MBZ-CASH", "MBZ", "Cash Box — MBZ", "Cash"],
        ["MUS-BANK", "Mussafah", "Main Bank — Mussafah", "Bank"],
        ["MUS-CASH", "Mussafah", "Cash Box — Mussafah", "Cash"],
        ["AF-BANK", "Al Falah", "ADIB BANK — Al Falah", "Bank"],
        ["AF-CASH", "Al Falah", "Cash Account — Al Falah", "Cash"],
    ]
    for index, values in enumerate(starter, 5):
        for col, value in enumerate(values, 1):
            ws.cell(index, col, value)
        ws.cell(index, 5, date.today())
        ws.cell(index, 6, 0)
        ws.cell(index, 8, "='Setup & Lists'!$M$6")
        ws.cell(index, 10, "Manual / POS")
        ws.cell(index, 5).comment = Comment("Replace with the latest reconciled bank or POS balance.", "Codex")
    for row in range(5, last_row + 1):
        ws.cell(row, 7, f'=IFERROR(IF(C{row}="","",SUMIFS(\'Cheque Register\'!$H$5:$H$304,\'Cheque Register\'!$D$5:$D$304,C{row},\'Cheque Register\'!$I$5:$I$304,"Pending")),0)')
        ws.cell(row, 9, f'=IFERROR(IF(F{row}="","",F{row}-G{row}-H{row}),0)')
        for col in [6, 7, 8, 9]:
            ws.cell(row, col).number_format = CURRENCY_FORMAT
        for col in [5, 11]:
            ws.cell(row, col).number_format = DATE_FORMAT
    add_table(ws, "tblBalances", f"A4:L{last_row}")
    add_list_validation(ws, f"B5:B{last_row}", "='Setup & Lists'!$A$5:$A$20")
    add_list_validation(ws, f"D5:D{last_row}", "='Setup & Lists'!$I$5:$I$10")
    ws.conditional_formatting.add(f"I5:I{last_row}", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor=RED), font=Font(color=DARK_RED)))
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    set_widths(ws, {"A": 16, "B": 14, "C": 28, "D": 16, "E": 13, "F": 20, "G": 20, "H": 22, "I": 22, "J": 18, "K": 18, "L": 28})
    return ws


def build_forecast(wb):
    ws = wb.create_sheet("Cash Flow Forecast")
    headers = ["Month", "Opening Balance (AED)", "Expected Income (AED)", "Open Commitments (AED)", "Fixed Templates (AED)", "Expected Expenses (AED)", "Net Cash Flow (AED)", "Closing Balance (AED)", "Liquidity Alert", "Notes"]
    title(ws, "12-Month Cash Flow Forecast", "دخل متوقع + التزامات فعلية + قوالب ثابتة، مع استبعاد Source Type = Fixed لمنع التكرار", len(headers))
    header_row(ws, 4, headers)
    last_row = 16
    style_input_area(ws, 5, last_row, [3, 10])
    style_formula_columns(ws, 5, last_row, [1, 2, 4, 5, 6, 7, 8, 9])
    for row in range(5, last_row + 1):
        month_offset = row - 5
        ws.cell(row, 1, f'=DATE(YEAR(TODAY()),MONTH(TODAY())+{month_offset},1)')
        ws.cell(row, 2, "=SUM('Bank & Cash Balances'!$I$5:$I$104)" if row == 5 else f"=H{row-1}")
        ws.cell(row, 4, f'=SUMIFS(Commitments!$L$5:$L$504,Commitments!$E$5:$E$504,">="&A{row},Commitments!$E$5:$E$504,"<="&EOMONTH(A{row},0),Commitments!$M$5:$M$504,"<>Paid",Commitments!$M$5:$M$504,"<>Cancelled",Commitments!$B$5:$B$504,"<>Fixed")')
        ws.cell(row, 5, f'=SUMIFS(\'Fixed Expenses\'!$E$5:$E$204,\'Fixed Expenses\'!$J$5:$J$204,"Yes",\'Fixed Expenses\'!$G$5:$G$204,"<="&EOMONTH(A{row},0),\'Fixed Expenses\'!$H$5:$H$204,">="&A{row})+SUMIFS(\'Fixed Expenses\'!$E$5:$E$204,\'Fixed Expenses\'!$J$5:$J$204,"Yes",\'Fixed Expenses\'!$G$5:$G$204,"<="&EOMONTH(A{row},0),\'Fixed Expenses\'!$H$5:$H$204,"")')
        ws.cell(row, 6, f'=D{row}+E{row}')
        ws.cell(row, 7, f'=C{row}-F{row}')
        ws.cell(row, 8, f'=B{row}+G{row}')
        ws.cell(row, 9, f'=IF(H{row}<0,"CASH DEFICIT",IF(H{row}<\'Setup & Lists\'!$M$6,"LOW LIQUIDITY","OK"))')
        ws.cell(row, 1).number_format = "mmm-yy"
        for col in range(2, 9):
            ws.cell(row, col).number_format = CURRENCY_FORMAT
    add_table(ws, "tblForecast", f"A4:J{last_row}")
    ws.conditional_formatting.add(f"H5:H{last_row}", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor=RED), font=Font(color=DARK_RED, bold=True)))
    ws.conditional_formatting.add(f"I5:I{last_row}", FormulaRule(formula=['$I5="CASH DEFICIT"'], fill=PatternFill("solid", fgColor=RED)))
    ws.conditional_formatting.add(f"I5:I{last_row}", FormulaRule(formula=['$I5="LOW LIQUIDITY"'], fill=PatternFill("solid", fgColor=ORANGE)))
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    set_widths(ws, {"A": 14, "B": 23, "C": 22, "D": 23, "E": 22, "F": 24, "G": 22, "H": 23, "I": 18, "J": 30})
    return ws


def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0)
    title(ws, "Multi-Branch Financial Dashboard", "لوحة موحدة لفروع MBZ وMussafah وAl Falah — القيم تتحدث من الصفحات المرتبطة", 12)
    ws["A4"] = "As of Date"
    ws["B4"] = "=TODAY()"
    ws["B4"].number_format = DATE_FORMAT
    ws["A4"].fill = PatternFill("solid", fgColor=GRAY)
    ws["B4"].fill = PatternFill("solid", fgColor=LIGHT_TEAL)

    kpi(ws, "A6", "A7", "Book Bank & Cash", "=SUM('Bank & Cash Balances'!$F$5:$F$104)")
    kpi(ws, "C6", "C7", "Available Liquidity", "=SUM('Bank & Cash Balances'!$I$5:$I$104)", fill=LIGHT_TEAL)
    kpi(ws, "E6", "E7", "Commitments — 30 Days", '=SUMIFS(Commitments!$L$5:$L$504,Commitments!$E$5:$E$504,">="&TODAY(),Commitments!$E$5:$E$504,"<="&TODAY()+30,Commitments!$M$5:$M$504,"<>Paid",Commitments!$M$5:$M$504,"<>Cancelled")', fill=ORANGE)
    kpi(ws, "G6", "G7", "Commitments — 60 Days", '=SUMIFS(Commitments!$L$5:$L$504,Commitments!$E$5:$E$504,">"&TODAY()+30,Commitments!$E$5:$E$504,"<="&TODAY()+60,Commitments!$M$5:$M$504,"<>Paid",Commitments!$M$5:$M$504,"<>Cancelled")', fill=YELLOW)
    kpi(ws, "I6", "I7", "Commitments — 90 Days", '=SUMIFS(Commitments!$L$5:$L$504,Commitments!$E$5:$E$504,">"&TODAY()+60,Commitments!$E$5:$E$504,"<="&TODAY()+90,Commitments!$M$5:$M$504,"<>Paid",Commitments!$M$5:$M$504,"<>Cancelled")', fill=YELLOW)
    kpi(ws, "K6", "K7", "Liquidity Coverage", '=IFERROR(C7/E7,0)', fill=LIGHT_TEAL, number_format="0.0x")

    kpi(ws, "A10", "A11", "Cheques Due This Month", '=SUMIFS(\'Cheque Register\'!$H$5:$H$304,\'Cheque Register\'!$G$5:$G$304,">="&EOMONTH(TODAY(),-1)+1,\'Cheque Register\'!$G$5:$G$304,"<="&EOMONTH(TODAY(),0),\'Cheque Register\'!$I$5:$I$304,"Pending")', fill=ORANGE)
    kpi(ws, "C10", "C11", "Salaries Due", '=SUMIFS(Commitments!$L$5:$L$504,Commitments!$F$5:$F$504,"Salaries",Commitments!$M$5:$M$504,"<>Paid",Commitments!$M$5:$M$504,"<>Cancelled")', fill=ORANGE)
    kpi(ws, "E10", "E11", "Unpaid Purchases", '=SUM(\'Accounts Payable\'!$K$5:$K$504)', fill=ORANGE)
    kpi(ws, "G10", "G11", "Overdue Suppliers", '=SUMIFS(\'Accounts Payable\'!$K$5:$K$504,\'Accounts Payable\'!$F$5:$F$504,"<"&TODAY(),\'Accounts Payable\'!$L$5:$L$504,"<>Paid",\'Accounts Payable\'!$L$5:$L$504,"<>Cancelled")', fill=RED)
    kpi(ws, "I10", "I11", "Open Commitments", '=SUMIFS(Commitments!$L$5:$L$504,Commitments!$M$5:$M$504,"<>Paid",Commitments!$M$5:$M$504,"<>Cancelled")', fill=YELLOW)
    kpi(ws, "K10", "K11", "Next Month Closing", "='Cash Flow Forecast'!H6", fill=LIGHT_TEAL)

    ws["A14"] = "12-Month Obligations"
    ws["A14"].font = Font(color=NAVY, bold=True, size=13)
    for index, row in enumerate(range(15, 27), 5):
        ws.cell(row, 1, f"='Cash Flow Forecast'!A{index}")
        ws.cell(row, 2, f"='Cash Flow Forecast'!F{index}")
        ws.cell(row, 3, f"='Cash Flow Forecast'!H{index}")
        ws.cell(row, 1).number_format = "mmm-yy"
        ws.cell(row, 2).number_format = CURRENCY_FORMAT
        ws.cell(row, 3).number_format = CURRENCY_FORMAT
    ws["A15"] = "='Cash Flow Forecast'!A5"
    ws["A14"].fill = PatternFill("solid", fgColor=LIGHT_TEAL)
    ws["B14"] = "Expected Expenses"
    ws["C14"] = "Closing Balance"
    for cell in ws[14]:
        if cell.column <= 3:
            cell.font = Font(color=NAVY, bold=True)
            cell.fill = PatternFill("solid", fgColor=LIGHT_TEAL)

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Monthly Obligations and Closing Liquidity"
    chart.y_axis.title = "AED"
    chart.x_axis.title = "Month"
    data = Reference(ws, min_col=2, max_col=3, min_row=14, max_row=26)
    cats = Reference(ws, min_col=1, min_row=15, max_row=26)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8.5
    chart.width = 16
    ws.add_chart(chart, "E14")

    ws["A29"] = "Alert Logic"
    ws["A29"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A29"].font = Font(color=WHITE, bold=True)
    alerts = [
        "Red: overdue supplier invoice, cheque due within 7 days, or negative forecast.",
        "Orange: payment due soon or liquidity below protected minimum.",
        "Source Type + Source ID must be unique before importing a commitment.",
        "Fixed templates feed the forecast; generated Fixed commitments are excluded from forecast commitments to prevent double counting.",
    ]
    for row, text in enumerate(alerts, 30):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        ws.cell(row, 1, text)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        ws.cell(row, 1).font = Font(color="555555")

    for row in [6, 7, 10, 11]:
        ws.row_dimensions[row].height = 28 if row in [6, 10] else 42
    ws.conditional_formatting.add("A7:K11", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor=RED)))
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    set_widths(ws, {"A": 20, "B": 18, "C": 20, "D": 4, "E": 20, "F": 4, "G": 20, "H": 4, "I": 20, "J": 4, "K": 20, "L": 4})
    return ws


def add_workbook_notes(wb):
    sources = {
        "Commitments": "Primary manual/import register. Future POS/n8n source: project workflows.",
        "Accounts Payable": "Expected source: POS purchases and supplier balances.",
        "Bank & Cash Balances": "Expected source: reconciled POS ledgers and bank statements.",
    }
    for sheet, note in sources.items():
        wb[sheet]["A2"].comment = Comment(note, "Codex")


def validate_workbook(path):
    workbook = load_workbook(path, data_only=False)
    required = ["Dashboard", "Commitments", "Cheque Register", "Accounts Payable", "Fixed Expenses", "Cash Flow Forecast", "Bank & Cash Balances", "Setup & Lists"]
    assert workbook.sheetnames == required, workbook.sheetnames
    assert workbook["Dashboard"]["E7"].value.startswith("=SUMIFS")
    assert workbook["Commitments"]["L5"].value == '=IF(I5="","",MAX(0,I5-K5))'
    assert workbook["Cash Flow Forecast"]["H5"].value == "=B5+G5"
    assert len(workbook["Dashboard"]._charts) == 1
    formula_errors = []
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("=") and "#REF!" in cell.value:
                    formula_errors.append(f"{ws.title}!{cell.coordinate}")
    assert not formula_errors, formula_errors
    return {"sheets": len(workbook.sheetnames), "formulas_checked": True, "charts": len(workbook["Dashboard"]._charts)}


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    build_setup(wb)
    build_commitments(wb)
    build_cheques(wb)
    build_ap(wb)
    build_fixed(wb)
    build_balances(wb)
    build_forecast(wb)
    build_dashboard(wb)
    add_workbook_notes(wb)
    order = ["Dashboard", "Commitments", "Cheque Register", "Accounts Payable", "Fixed Expenses", "Cash Flow Forecast", "Bank & Cash Balances", "Setup & Lists"]
    wb._sheets = [wb[name] for name in order]
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(OUTPUT_FILE)
    result = validate_workbook(OUTPUT_FILE)
    print(OUTPUT_FILE)
    print(result)


if __name__ == "__main__":
    main()
